#!/usr/bin/env python3
"""Sanitize the raw Porsche sales dataset per SchemaPorshe.md.

Reads the raw workbook (data/porsche_sales_raw.xlsx) and writes a canonical
CSV (data/porsche_sales_sanitized.csv) with 22 columns: each raw column is
followed immediately by its sanitized counterpart, per the schema's Quality
checks ("No original columns were removed", "Sanitized columns appear
immediately after their source columns").
"""
import csv
import datetime
import re
import sys

import openpyxl

RAW_PATH = "data/porsche_sales_raw.xlsx"
OUT_PATH = "data/porsche_sales_sanitized.csv"

MIN_YEAR, MAX_YEAR = 1990, 2035
KM_TO_MILES = 0.621371

WORDS_ONES = {
    "zero": 0, "one": 1, "two": 2, "three": 3, "four": 4, "five": 5, "six": 6,
    "seven": 7, "eight": 8, "nine": 9, "ten": 10, "eleven": 11, "twelve": 12,
    "thirteen": 13, "fourteen": 14, "fifteen": 15, "sixteen": 16,
    "seventeen": 17, "eighteen": 18, "nineteen": 19,
}
WORDS_TENS = {
    "twenty": 20, "thirty": 30, "forty": 40, "fifty": 50, "sixty": 60,
    "seventy": 70, "eighty": 80, "ninety": 90,
}
MONTHS = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11,
    "december": 12, "jan": 1, "feb": 2, "mar": 3, "apr": 4, "jun": 6, "jul": 7,
    "aug": 8, "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
}
US_STATES = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR",
    "california": "CA", "colorado": "CO", "connecticut": "CT", "delaware": "DE",
    "florida": "FL", "georgia": "GA", "hawaii": "HI", "idaho": "ID",
    "illinois": "IL", "indiana": "IN", "iowa": "IA", "kansas": "KS",
    "kentucky": "KY", "louisiana": "LA", "maine": "ME", "maryland": "MD",
    "massachusetts": "MA", "michigan": "MI", "minnesota": "MN",
    "mississippi": "MS", "missouri": "MO", "montana": "MT", "nebraska": "NE",
    "nevada": "NV", "new hampshire": "NH", "new jersey": "NJ",
    "new mexico": "NM", "new york": "NY", "north carolina": "NC",
    "north dakota": "ND", "ohio": "OH", "oklahoma": "OK", "oregon": "OR",
    "pennsylvania": "PA", "rhode island": "RI", "south carolina": "SC",
    "south dakota": "SD", "tennessee": "TN", "texas": "TX", "utah": "UT",
    "vermont": "VT", "virginia": "VA", "washington": "WA",
    "west virginia": "WV", "wisconsin": "WI", "wyoming": "WY",
    "district of columbia": "DC",
}
VALID_CODES = set(US_STATES.values())

PAYMENT_RULES = [
    ("ach", "ACH Payment"),
    ("crypto", "Crypto Payment"),
    ("wire", "Wire Transfer"),
    ("bank", "Bank Transfer"),
    ("debit", "Debit Card"),
    ("credit", "Credit Card"),
    ("financ", "Financing"),
    ("leas", "Lease"),
    ("cash", "Cash"),
]

DELIVERY_MAP = {
    "delivered": "Delivered",
    "deliverd": "Delivered",
    "pending": "Pending",
    "in transit": "In Transit",
    "cancelled": "Cancelled",
    "canceled": "Cancelled",
    "awaiting delivery": "Awaiting Delivery",
    "awaiting pickup": "Awaiting Pickup",
    "pending approval": "Pending Approval",
    "pending review": "Pending Review",
    "shipped": "Shipped",
    "awaiting review": "Awaiting Review",
}


def parse_cardinal(tokens):
    """Standard English cardinal number parsing: ones/tens + hundred/thousand."""
    total = 0
    current = 0
    for w in tokens:
        if w == "hundred":
            current = (current or 1) * 100
        elif w == "thousand":
            total += (current or 1) * 1000
            current = 0
        elif w in WORDS_ONES:
            current += WORDS_ONES[w]
        elif w in WORDS_TENS:
            current += WORDS_TENS[w]
        else:
            return None
    return total + current


def parse_number_grouping(s):
    """Parse a numeric string with ambiguous ',' / '.' as thousands or decimal sep."""
    has_comma, has_period = "," in s, "." in s
    if has_comma and has_period:
        if s.rfind(",") > s.rfind("."):
            s2 = s.replace(".", "").replace(",", ".")
        else:
            s2 = s.replace(",", "")
        return float(s2)
    if has_comma:
        idx = s.rfind(",")
        frac_len = len(s) - idx - 1
        return float(s.replace(",", "")) if frac_len == 3 else float(s.replace(",", "."))
    if has_period:
        idx = s.rfind(".")
        frac_len = len(s) - idx - 1
        return float(s.replace(".", "")) if frac_len == 3 else float(s)
    return float(s)


def sanitize_date(raw):
    if raw is None or str(raw).strip() == "":
        return "INVALID"
    if isinstance(raw, (datetime.datetime, datetime.date)):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()

    def valid(y, mo, d):
        try:
            datetime.date(y, mo, d)
            return True
        except ValueError:
            return False

    m = re.fullmatch(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y:04d}-{mo:02d}-{d:02d}" if valid(y, mo, d) else "INVALID"

    m = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", s)
    if m:
        mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y:04d}-{mo:02d}-{d:02d}" if valid(y, mo, d) else "INVALID"

    m = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2})", s)
    if m:
        mo, d, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        y = 2000 + yy
        return f"{y:04d}-{mo:02d}-{d:02d}" if valid(y, mo, d) else "INVALID"

    m = re.fullmatch(r"([A-Za-z]+)\.?\s+(\d{1,2})(?:st|nd|rd|th)?,?\s+(\d{4})", s)
    if m:
        mo = MONTHS.get(m.group(1).lower())
        d, y = int(m.group(2)), int(m.group(3))
        if mo is None:
            return "INVALID"
        return f"{y:04d}-{mo:02d}-{d:02d}" if valid(y, mo, d) else "INVALID"

    return "INVALID"


def sanitize_model(raw):
    if raw is None or str(raw).strip() == "":
        return "INVALID"
    s = str(raw).strip().title()
    return re.sub(r"\b(gt3|gt4|gts|rs|gt)\b", lambda m: m.group(0).upper(), s, flags=re.IGNORECASE)


def sanitize_year(raw):
    if raw is None or str(raw).strip() == "":
        return "INVALID"
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        y = int(raw)
        return str(y) if MIN_YEAR <= y <= MAX_YEAR else "INVALID"
    s = str(raw).strip().lower()
    if s.isdigit():
        y = int(s)
        return str(y) if MIN_YEAR <= y <= MAX_YEAR else "INVALID"

    m = re.fullmatch(r"(\d{2})[\s-](\d{2})", s)
    if m:
        y = int(m.group(1)) * 100 + int(m.group(2))
        return str(y) if MIN_YEAR <= y <= MAX_YEAR else "INVALID"

    tokens = re.findall(r"[a-z]+", s)
    if "thousand" in tokens:
        y = parse_cardinal(tokens)
        return str(y) if y and MIN_YEAR <= y <= MAX_YEAR else "INVALID"

    # spoken digit-pair style: "twenty twenty four" -> 20|24
    for split in range(1, len(tokens)):
        left, right = tokens[:split], tokens[split:]
        lv = parse_cardinal(left) if len(left) == 1 else None
        rv = parse_cardinal(right)
        if lv is not None and 0 <= lv <= 99 and rv is not None and 0 <= rv <= 99 and len(right) <= 2:
            y = lv * 100 + rv
            if MIN_YEAR <= y <= MAX_YEAR:
                return str(y)
    return "INVALID"


def sanitize_price(raw):
    if raw is None or str(raw).strip() == "":
        return "INVALID"
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return f"{float(raw):.2f}"
    s = str(raw).strip()
    low = s.lower()
    tokens = re.findall(r"[a-z]+", low)
    word_tokens = [t for t in tokens if t not in ("usd", "dollars", "dollar")]
    if word_tokens and all(t in WORDS_ONES or t in WORDS_TENS or t in ("hundred", "thousand") for t in word_tokens):
        val = parse_cardinal(word_tokens)
        if val is not None:
            return f"{float(val):.2f}"

    cleaned = re.sub(r"(?i)usd|dollars?|\$", "", s).strip()
    is_k = bool(re.search(r"[kK]\s*$", cleaned))
    cleaned = re.sub(r"[kK]\s*$", "", cleaned).strip()
    cleaned = cleaned.replace(" ", "")
    if not cleaned:
        return "INVALID"
    try:
        val = parse_number_grouping(cleaned)
    except ValueError:
        return "INVALID"
    if is_k:
        val *= 1000
    return f"{val:.2f}"


def sanitize_mileage(raw):
    if raw is None or str(raw).strip() == "":
        return "INVALID"
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return str(round(float(raw)))
    s = str(raw).strip()
    low = s.lower()

    if low in ("new", "new car", "zero", "zero miles", "0 mi", "0 miles"):
        return "0"

    tokens = re.findall(r"[a-z]+", low)
    word_tokens = [t for t in tokens if t not in ("miles", "mi", "km")]
    if word_tokens and all(t in WORDS_ONES or t in WORDS_TENS or t in ("hundred", "thousand") for t in word_tokens):
        val = parse_cardinal(word_tokens)
        if val is not None:
            return str(round(val))

    is_km = bool(re.search(r"\bkm\b", low))
    cleaned = re.sub(r"(?i)(miles?|mi\.?|km)\b", "", s)
    cleaned = re.sub(r"[:\s]", "", cleaned)
    if not cleaned:
        return "INVALID"
    try:
        val = parse_number_grouping(cleaned)
    except ValueError:
        return "INVALID"
    if is_km:
        val = val * KM_TO_MILES
    return str(round(val))


def sanitize_payment(raw):
    if raw is None or str(raw).strip() == "":
        return "INVALID"
    norm = re.sub(r"[-_]", " ", str(raw).strip().lower())
    for keyword, label in PAYMENT_RULES:
        if keyword in norm:
            return label
    return str(raw).strip().title()


def sanitize_city(raw):
    if raw is None or str(raw).strip() == "":
        return "INVALID"
    s = str(raw).strip()
    return s.title() if s == s.lower() else s


def sanitize_state(raw):
    if raw is None or str(raw).strip() == "":
        return "INVALID"
    s = str(raw).strip()
    if s.upper() in VALID_CODES:
        return s.upper()
    return US_STATES.get(s.lower(), "INVALID")


def sanitize_delivery(raw):
    if raw is None or str(raw).strip() == "":
        return "INVALID"
    norm = re.sub(r"[^a-z\s]", " ", str(raw).strip().lower())
    norm = re.sub(r"[-]", " ", norm)
    norm = re.sub(r"\s+", " ", norm).strip()
    return DELIVERY_MAP.get(norm, "INVALID")


def main():
    wb = openpyxl.load_workbook(RAW_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    header = [
        "sale_id",
        "sale_date", "SaleDateSanitized",
        "customer_name",
        "porsche_model", "PorscheModelSanitized",
        "model_year", "ModelYearSanitized",
        "sale_price", "SalesPriceSanitized",
        "vehicle_mileage", "VehicleMileageSanitized",
        "payment_method", "PayMethodSanitized",
        "city", "CitySanitized",
        "state", "StateSanitized",
        "salesperson",
        "delivery_status", "DeliveryStatusSanitized",
    ]

    out_rows = []
    for r in rows:
        (sale_id, sale_date, customer_name, porsche_model, model_year, sale_price,
         vehicle_mileage, payment_method, city, state, salesperson, delivery_status) = r

        raw_date = sale_date.strftime("%Y-%m-%d") if isinstance(sale_date, (datetime.datetime, datetime.date)) else sale_date

        out_rows.append([
            sale_id,
            raw_date, sanitize_date(sale_date),
            customer_name,
            porsche_model, sanitize_model(porsche_model),
            model_year, sanitize_year(model_year),
            sale_price, sanitize_price(sale_price),
            vehicle_mileage, sanitize_mileage(vehicle_mileage),
            payment_method, sanitize_payment(payment_method),
            city, sanitize_city(city),
            state, sanitize_state(state),
            salesperson,
            delivery_status, sanitize_delivery(delivery_status),
        ])

    with open(OUT_PATH, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(out_rows)

    print(f"Wrote {len(out_rows)} rows to {OUT_PATH}", file=sys.stderr)


if __name__ == "__main__":
    main()
